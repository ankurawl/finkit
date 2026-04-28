"""CSV/XLS/XLSX importer with auto-detection, two-phase mapping, and deduplication."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

from personalfinance.config import get_config, resolve_path
from personalfinance.ledger import append_text, format_transaction, get_accounts, load_file
from personalfinance.uuids import generate_uuid_tag


DATE_FORMATS = [
    "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
    "%Y/%m/%d", "%m-%d-%Y", "%d-%m-%Y", "%b %d, %Y", "%B %d, %Y",
]

DATE_COLUMN_NAMES = {"date", "posting date", "trans date", "transaction date", "trade date", "settlement date", "post date"}
AMOUNT_COLUMN_NAMES = {"amount", "debit", "credit", "total", "net amount", "transaction amount"}
PAYEE_COLUMN_NAMES = {"description", "payee", "merchant", "name", "memo", "narrative", "details", "transaction description"}
SKIP_COLUMN_NAMES = {"balance", "running balance", "running bal", "running bal.", "type", "category", "check number", "reference"}


def import_file(
    file_path: str,
    account: str,
    mapping_name: str | None = None,
    confirm_mapping: dict | None = None,
    sheet_name: str | None = None,
    ledger_path: str | None = None,
) -> dict[str, Any]:
    """
    Import transactions from CSV/XLS/XLSX.

    Phase 1 (no confirm_mapping): detect format and return proposed mapping.
    Phase 2 (with confirm_mapping): apply mapping, import, deduplicate.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        headers, rows = _read_csv(path)
    elif ext == ".xls":
        headers, rows, sheets = _read_xls(path, sheet_name)
        if sheets and not sheet_name and len(sheets) > 1:
            return {
                "status": "select_sheet",
                "sheets": sheets,
                "message": "Multiple sheets found. Specify sheet_name.",
            }
    elif ext in (".xlsx", ".xlsm"):
        headers, rows, sheets = _read_xlsx(path, sheet_name)
        if sheets and not sheet_name and len(sheets) > 1:
            return {
                "status": "select_sheet",
                "sheets": sheets,
                "message": "Multiple sheets found. Specify sheet_name.",
            }
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use CSV, XLS, or XLSX.")

    if mapping_name and not confirm_mapping:
        saved = _load_mapping(mapping_name)
        if saved:
            confirm_mapping = saved

    if confirm_mapping is None:
        mapping = _detect_columns(headers, rows)
        return {
            "status": "mapping_proposed",
            "file": str(path),
            "headers": headers,
            "detected_mapping": mapping,
            "sample_rows": [dict(zip(headers, row)) for row in rows[:3]],
            "row_count": len(rows),
            "message": "Review the detected mapping. To proceed, call again with confirm_mapping.",
        }

    transactions = _parse_transactions(headers, rows, confirm_mapping, account)

    if ledger_path is None:
        from personalfinance.config import get_ledger_path
        ledger_path = str(get_ledger_path())

    new_txns, dupes = _deduplicate(transactions, ledger_path)

    for txn in new_txns:
        text = format_transaction(
            date_=txn["date"],
            payee=txn.get("payee"),
            narration=txn.get("narration", "Imported transaction"),
            postings=txn["postings"],
            tags={txn["uuid_tag"]},
        )
        append_text(ledger_path, text)

    if mapping_name or confirm_mapping.get("save_as"):
        name = mapping_name or confirm_mapping["save_as"]
        _save_mapping(name, confirm_mapping)

    date_range = ""
    if new_txns:
        dates = [t["date"] for t in new_txns]
        date_range = f"{min(dates).isoformat()} to {max(dates).isoformat()}"

    return {
        "status": "imported",
        "imported_count": len(new_txns),
        "duplicate_count": len(dupes),
        "date_range": date_range,
        "account": account,
        "mapping_saved": mapping_name or confirm_mapping.get("save_as"),
    }


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    """Read a CSV file with auto-detected delimiter."""
    content = path.read_text(encoding="utf-8-sig")
    dialect = csv.Sniffer().sniff(content[:4096])
    reader = csv.reader(io.StringIO(content), dialect)
    rows_raw = list(reader)
    if not rows_raw:
        return [], []
    headers = [h.strip() for h in rows_raw[0]]
    rows = [r for r in rows_raw[1:] if any(cell.strip() for cell in r)]
    return headers, rows


def _read_xlsx(path: Path, sheet_name: str | None) -> tuple[list[str], list[list[str]], list[str]]:
    """Read an XLSX file."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = wb.sheetnames

    target = sheet_name or sheets[0]
    ws = wb[target]

    rows_raw = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()

    if not rows_raw:
        return [], [], sheets
    headers = [h.strip() for h in rows_raw[0]]
    rows = [r for r in rows_raw[1:] if any(cell.strip() for cell in r)]
    return headers, rows, sheets


def _read_xls(path: Path, sheet_name: str | None) -> tuple[list[str], list[list[str]], list[str]]:
    """Read an XLS (legacy Excel) file."""
    import xlrd
    wb = xlrd.open_workbook(str(path))
    sheets = wb.sheet_names()

    target = sheet_name or sheets[0]
    ws = wb.sheet_by_name(target)

    if ws.nrows == 0:
        return [], [], sheets
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    rows = []
    for r in range(1, ws.nrows):
        row = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
        if any(cell.strip() for cell in row):
            rows.append(row)
    return headers, rows, sheets


def _detect_columns(headers: list[str], rows: list[list[str]]) -> dict[str, Any]:
    """Auto-detect column roles based on header names and data patterns."""
    mapping: dict[str, Any] = {}
    headers_lower = [h.lower().strip() for h in headers]

    for i, h in enumerate(headers_lower):
        if h in DATE_COLUMN_NAMES:
            mapping["date_col"] = headers[i]
            if rows:
                mapping["date_format"] = _detect_date_format(rows[0][i])
        elif h in AMOUNT_COLUMN_NAMES:
            if "amount_col" not in mapping:
                mapping["amount_col"] = headers[i]
                mapping["amount_negate"] = _detect_amount_negate(h, [r[i] for r in rows[:10]])
        elif h in PAYEE_COLUMN_NAMES:
            mapping["payee_col"] = headers[i]
        elif h in SKIP_COLUMN_NAMES:
            mapping.setdefault("skip_cols", []).append(headers[i])

    if "debit" in headers_lower and "credit" in headers_lower:
        mapping["debit_col"] = headers[headers_lower.index("debit")]
        mapping["credit_col"] = headers[headers_lower.index("credit")]
        mapping.pop("amount_col", None)

    if "date_col" not in mapping:
        for i, h in enumerate(headers_lower):
            if rows and _detect_date_format([r[i] for r in rows[:5] if r[i].strip()][0] if any(r[i].strip() for r in rows[:5]) else ""):
                mapping["date_col"] = headers[i]
                mapping["date_format"] = _detect_date_format(rows[0][i])
                break

    return mapping


def _detect_date_format(sample: str) -> str | None:
    """Try parsing a date string with common formats."""
    sample = sample.strip()
    for fmt in DATE_FORMATS:
        try:
            datetime.strptime(sample, fmt)
            return fmt
        except ValueError:
            continue
    return None


def _detect_amount_negate(col_name: str, samples: list[str]) -> bool:
    """Detect if amounts should be negated (e.g., negative = debit in some bank CSVs)."""
    return False


def _parse_transactions(
    headers: list[str],
    rows: list[list[str]],
    mapping: dict,
    account: str,
) -> list[dict]:
    """Parse rows into transaction dicts using the confirmed mapping."""
    transactions = []
    date_col = mapping.get("date_col")
    amount_col = mapping.get("amount_col")
    debit_col = mapping.get("debit_col")
    credit_col = mapping.get("credit_col")
    payee_col = mapping.get("payee_col")
    date_format = mapping.get("date_format", "%m/%d/%Y")
    negate = mapping.get("amount_negate", False)

    header_idx = {h: i for i, h in enumerate(headers)}

    for row in rows:
        if len(row) < len(headers):
            row.extend([""] * (len(headers) - len(row)))

        date_str = row[header_idx[date_col]].strip() if date_col and date_col in header_idx else ""
        if not date_str:
            continue

        try:
            txn_date = datetime.strptime(date_str, date_format).date()
        except (ValueError, TypeError):
            for fmt in DATE_FORMATS:
                try:
                    txn_date = datetime.strptime(date_str, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                continue

        if debit_col and credit_col:
            debit_str = row[header_idx.get(debit_col, 0)].strip() if debit_col in header_idx else ""
            credit_str = row[header_idx.get(credit_col, 0)].strip() if credit_col in header_idx else ""
            debit_val = _parse_amount(debit_str)
            credit_val = _parse_amount(credit_str)
            if debit_val:
                amount = -abs(debit_val)
            elif credit_val:
                amount = abs(credit_val)
            else:
                continue
        elif amount_col and amount_col in header_idx:
            amount_str = row[header_idx[amount_col]].strip()
            amount = _parse_amount(amount_str)
            if amount is None:
                continue
            if negate:
                amount = -amount
        else:
            continue

        payee = row[header_idx[payee_col]].strip() if payee_col and payee_col in header_idx else None

        from personalfinance.config import get_config
        currency = get_config().general.default_currency

        postings = [
            {"account": account, "amount": str(amount), "currency": currency},
            {"account": "Expenses:Other" if amount < 0 else "Income:Other", "amount": str(-amount), "currency": currency},
        ]

        transactions.append({
            "date": txn_date,
            "payee": payee,
            "narration": payee or "Imported transaction",
            "postings": postings,
            "uuid_tag": generate_uuid_tag(),
            "_hash": _txn_hash(txn_date, amount, payee or ""),
        })

    return transactions


def _parse_amount(s: str) -> Decimal | None:
    """Parse an amount string, handling currency symbols, commas, parens."""
    if not s:
        return None
    s = s.strip()
    s = re.sub(r'[$€£¥₹,]', '', s)
    negative = False
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1]
        negative = True
    try:
        val = Decimal(s)
        return -val if negative else val
    except InvalidOperation:
        return None


def _txn_hash(txn_date: date, amount: Decimal, description: str) -> str:
    """Create a dedup hash from date + amount + description."""
    key = f"{txn_date.isoformat()}|{amount}|{description.lower().strip()}"
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def _deduplicate(
    new_transactions: list[dict],
    ledger_path: str,
) -> tuple[list[dict], list[dict]]:
    """Remove transactions that already exist in the ledger."""
    try:
        entries, _, _ = load_file(ledger_path)
    except FileNotFoundError:
        return new_transactions, []

    from beancount.core import data as bc_data
    existing_hashes = set()
    config = get_config()
    window = config.import_.dedup_window_days

    for entry in entries:
        if isinstance(entry, bc_data.Transaction):
            for p in entry.postings:
                if p.units and p.units.number:
                    h = _txn_hash(entry.date, p.units.number, entry.payee or entry.narration or "")
                    existing_hashes.add(h)

    unique = []
    dupes = []
    for txn in new_transactions:
        if txn["_hash"] in existing_hashes:
            dupes.append(txn)
        else:
            unique.append(txn)

    return unique, dupes


def _load_mapping(name: str) -> dict | None:
    """Load a saved column mapping by name."""
    mappings_dir = resolve_path(get_config().import_.mappings_dir)
    mapping_file = mappings_dir / f"{name}.json"
    if mapping_file.exists():
        return json.loads(mapping_file.read_text())
    return None


def _save_mapping(name: str, mapping: dict) -> None:
    """Save a column mapping for future reuse."""
    mappings_dir = resolve_path(get_config().import_.mappings_dir)
    mappings_dir.mkdir(parents=True, exist_ok=True)
    mapping_file = mappings_dir / f"{name}.json"
    clean = {k: v for k, v in mapping.items() if k != "save_as"}
    mapping_file.write_text(json.dumps(clean, indent=2))
